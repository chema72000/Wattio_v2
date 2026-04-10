"""Tool for searching the thermal dissipation capacity database to pre-select cores."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from wattio.models import ToolResult
from wattio.tools.base import BaseTool

# Families to consider for PSFB transformer core selection
TARGET_FAMILIES = {"E", "PQ", "ETD", "RM"}

# Mandatory families — at least one core from each must appear in results
MANDATORY_FAMILIES = {"E", "PQ"}

# Map user-friendly cooling names to convection_velocity values in the database
COOLING_MAP = {
    "natural": 0,
    "free": 0,
    "forced_1": 1,
    "forced_4": 4,
}

EXCEL_FILENAME = "magnetic_core_thermal-2.xlsx"

# Loss distribution labels
PCT_LABELS = {10: "winding-heavy", 50: "balanced", 90: "core-heavy"}

# Category labels
CATEGORY_LABELS = {
    "aggressive": "aggressive (capacity < losses, needs optimization)",
    "middle": "middle (capacity near losses)",
    "conservative": "conservative (30%+ margin)",
}


def _parse_family(core_name: str) -> str:
    """Extract family prefix from core name (e.g., 'ETD34/17/11' -> 'ETD')."""
    family = ""
    for ch in core_name:
        if ch.isalpha() or ch == "_":
            family += ch
        else:
            break
    return family


def _find_excel(project_dir: Path) -> Path:
    """Locate the thermal database Excel file."""
    for candidate in [
        project_dir / EXCEL_FILENAME,
        project_dir.parent / EXCEL_FILENAME,
        Path(__file__).parent.parent.parent.parent / EXCEL_FILENAME,
    ]:
        if candidate.is_file():
            return candidate
    raise FileNotFoundError(
        f"Cannot find {EXCEL_FILENAME}. Place it in the project directory."
    )


def _load_and_search(
    project_dir: Path,
    total_losses: float,
    convection_velocity: int,
) -> list[dict]:
    """Load Excel and find 6 cores across families with 3 tiers.

    Returns 6 cores:
    - 2 aggressive: best_total_capacity at 90% of total_losses (10% below)
    - 2 middle: best_total_capacity just above total_losses (100-130%)
    - 2 conservative: best_total_capacity >= 130% of total_losses

    Ensures at least one E, one PQ, and one planar core in the results.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required. Install with: uv add openpyxl")

    excel_path = _find_excel(project_dir)
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb.active

    # Parse header
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(headers)}

    # Collect ALL rows for the cooling condition
    all_rows: dict[tuple[str, int], dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col["n_stacks"]] != 1:
            continue
        if row[col["convection_velocity"]] != convection_velocity:
            continue

        core_name = row[col["core_name"]]
        family = _parse_family(core_name)
        if family not in TARGET_FAMILIES:
            continue

        p_pct = row[col["P_%_ferrite"]]
        p_core = row[col["p_core"]] or 0
        p_winding = row[col["p_winding"]] or 0
        total_capacity = p_core + p_winding
        eff_volume = row[col["effective_volume"]] or 0
        eff_area = row[col["effective_area"]] or 0

        all_rows[(core_name, p_pct)] = {
            "core_name": core_name,
            "family": family,
            "p_pct_ferrite": p_pct,
            "p_core": round(p_core, 3),
            "p_winding": round(p_winding, 3),
            "total_capacity": round(total_capacity, 3),
            "effective_volume": eff_volume,
            "effective_area": eff_area,
        }

    wb.close()

    # Group by core_name with all 3 distributions
    cores: dict[str, dict] = {}
    for (core_name, p_pct), data in all_rows.items():
        if core_name not in cores:
            cores[core_name] = {
                "core_name": core_name,
                "distributions": {},
                "family": data["family"],
                "effective_volume": data["effective_volume"],
            }
        cores[core_name]["distributions"][p_pct] = data

    # Build candidates with best_total_capacity
    candidates: list[dict] = []
    for core_name, info in cores.items():
        dists = info["distributions"]
        if not all(p in dists for p in (10, 50, 90)):
            continue

        best_pct = max((10, 50, 90), key=lambda p: dists[p]["total_capacity"])
        best_total = dists[best_pct]["total_capacity"]

        candidates.append({
            "core_name": core_name,
            "family": info["family"],
            "effective_volume": info["effective_volume"],
            "dist_10": dists[10],
            "dist_50": dists[50],
            "dist_90": dists[90],
            "best_distribution": best_pct,
            "best_total_capacity": best_total,
        })

    # Sort by effective_volume (smallest first)
    candidates.sort(key=lambda c: c["effective_volume"])

    # Categorize cores into 3 tiers based on best_total_capacity vs total_losses
    aggressive_threshold = total_losses * 0.90   # capacity < losses (10% below)
    conservative_threshold = total_losses * 1.30  # capacity >= 130% of losses

    aggressive: list[dict] = []  # best_total < total_losses but >= 90%
    middle: list[dict] = []      # total_losses <= best_total < 130%
    conservative: list[dict] = []  # best_total >= 130%

    for c in candidates:
        cap = c["best_total_capacity"]
        if cap < aggressive_threshold:
            continue  # Too small, skip
        elif cap < total_losses:
            c["category"] = "aggressive"
            aggressive.append(c)
        elif cap < conservative_threshold:
            c["category"] = "middle"
            middle.append(c)
        else:
            c["category"] = "conservative"
            conservative.append(c)

    # Select 6 cores: 2 aggressive, 2 middle, 2 conservative
    # Ensure at least one E, one PQ, and one planar
    def _select_from_tier(tier: list[dict], count: int) -> list[dict]:
        """Pick smallest cores from tier, up to count."""
        # Already sorted by volume (smallest first)
        return tier[:count]

    selected: list[dict] = []
    selected.extend(_select_from_tier(aggressive, 2))
    selected.extend(_select_from_tier(middle, 2))
    selected.extend(_select_from_tier(conservative, 2))

    # Check mandatory families are represented
    families_present = {c["family"] for c in selected}

    all_candidates = aggressive + middle + conservative
    for fam in MANDATORY_FAMILIES:
        if fam not in families_present:
            for c in all_candidates:
                if c not in selected and c["family"] == fam:
                    selected.append(c)
                    break

    # Sort final selection: aggressive first (smallest), then middle, then conservative
    tier_order = {"aggressive": 0, "middle": 1, "conservative": 2}
    selected.sort(key=lambda c: (tier_order.get(c.get("category", "middle"), 1), c["effective_volume"]))

    return selected


class CoreThermalSearchTool(BaseTool):
    """Search the thermal dissipation database for viable transformer cores."""

    name = "core_thermal_search"
    description = (
        "Search the thermal dissipation capacity database (Excel) to pre-select "
        "transformer cores BEFORE using the Frenetic Core Optimizer. "
        "Returns 6 cores across E, PQ, ETD, RM families in 3 tiers: "
        "2 aggressive (capacity 10% below losses), 2 middle (capacity near losses), "
        "2 conservative (30%+ margin). Shows ALL 3 loss distribution scenarios "
        "(core-heavy, balanced, winding-heavy) for each core. "
        "Use this after calculating the loss budget (Step 7) and before the Core Optimizer."
    )
    parameters = {
        "type": "object",
        "properties": {
            "total_losses": {
                "type": "number",
                "description": "Expected total transformer losses in watts (core + winding).",
            },
            "cooling": {
                "type": "string",
                "enum": ["natural", "forced_1", "forced_4"],
                "description": "Cooling condition: 'natural' (free convection), 'forced_1' (1 m/s), 'forced_4' (4 m/s).",
            },
        },
        "required": ["total_losses", "cooling"],
    }

    async def execute(self, project_dir: Path, **kwargs: Any) -> ToolResult:
        total_losses = kwargs.get("total_losses")
        cooling = kwargs.get("cooling", "natural")

        if not total_losses or total_losses <= 0:
            return ToolResult(
                tool_call_id="",
                content="Error: total_losses must be a positive number.",
                is_error=True,
            )

        convection_velocity = COOLING_MAP.get(cooling, 0)

        try:
            results = _load_and_search(
                project_dir=project_dir,
                total_losses=total_losses,
                convection_velocity=convection_velocity,
            )
        except FileNotFoundError as e:
            return ToolResult(tool_call_id="", content=f"Error: {e}", is_error=True)
        except Exception as e:
            return ToolResult(
                tool_call_id="", content=f"Error searching database: {e}", is_error=True
            )

        if not results:
            return ToolResult(
                tool_call_id="",
                content=(
                    f"No cores found with total dissipation capacity >= {total_losses * 0.9:.2f} W "
                    f"(90% of {total_losses:.2f} W) under {cooling} convection. "
                    f"Consider: (1) forced convection, (2) larger ΔT, or (3) reducing loss target."
                ),
            )

        # Format results
        lines = [
            f"## Core Pre-Selection Results",
            f"**Target:** total transformer losses = {total_losses:.2f} W",
            f"**Conditions:** {cooling} convection, n_stacks = 1",
            f"**Database:** t_amb = 25°C, t_max = 100°C (ΔT = 75°C)",
            f"**⚠️ IMPORTANT:** These capacity values are the losses at ΔT = 75°C (100°C max). "
            f"This is a REFERENCE POINT, not a hard limit. The core can dissipate MORE losses — "
            f"it will simply run hotter. The real limit is the material's max operating temperature "
            f"(typically 120°C for most ferrites, higher for some). Do NOT say a core 'can't handle' "
            f"losses above its table capacity — it can, it will just be hotter than 100°C.",
            f"**Tiers:** aggressive (<100% capacity), middle (100-130%), conservative (>130%)",
            "",
        ]

        for i, c in enumerate(results, 1):
            vol_mm3 = c["effective_volume"] * 1e9  # m³ to mm³
            d10 = c["dist_10"]
            d50 = c["dist_50"]
            d90 = c["dist_90"]
            best = c["best_distribution"]
            category = c.get("category", "middle")
            cat_label = CATEGORY_LABELS.get(category, category)

            lines.append(
                f"### {i}. {c['core_name']} ({c['family']}) — {cat_label} — "
                f"Eff. Volume: {vol_mm3:.0f} mm³"
            )
            lines.append("")
            lines.append("| Distribution | P_core (W) | P_winding (W) | Total capacity (W) |")
            lines.append("|-------------|-----------|--------------|-------------------|")
            for pct, dist in [(10, d10), (50, d50), (90, d90)]:
                marker = " **<- best**" if pct == best else ""
                lines.append(
                    f"| {PCT_LABELS[pct]} ({pct}% ferrite) | "
                    f"{dist['p_core']:.3f} | {dist['p_winding']:.3f} | "
                    f"{dist['total_capacity']:.3f}{marker} |"
                )
            lines.append("")
            lines.append(
                f"**Optimal strategy:** Target **{PCT_LABELS[best]}** loss distribution "
                f"({best}% ferrite) for maximum thermal capacity of {c['best_total_capacity']:.3f} W."
            )
            lines.append("")

        lines.append(
            "**Testing order:** Start from the most aggressive (smallest) core. "
            "For each core, open the Core tab, select the core and material, "
            "then use the inductance calculator to find valid turns/inductance/gap "
            "combinations before generating a winding."
        )

        return ToolResult(
            tool_call_id="",
            content="\n".join(lines),
        )
