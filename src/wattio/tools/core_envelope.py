"""Tool for looking up the outer bounding-box (envelope) of catalog cores.

Reads dimensions from ``src/wattio/data/core_geometry_db.xlsx`` and returns
the three outer dimensions of each pair-stacked core (the physical bounding
box). Dimensions are returned **sorted descending**, so comparing against a
user envelope is orientation-agnostic: sort the user's constraints the same
way and compare element-wise.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from wattio.models import ToolResult
from wattio.tools.base import BaseTool


DB_FILENAME = "core_geometry_db.xlsx"


def _find_db() -> Path:
    """Locate the core geometry Excel DB shipped with the package."""
    here = Path(__file__).resolve().parent.parent / "data" / DB_FILENAME
    if here.is_file():
        return here
    raise FileNotFoundError(f"Cannot find {DB_FILENAME} in package data dir.")


def _envelope_for_row(family: str, headers: list[str], row: tuple) -> tuple[float, float, float] | None:
    """Return (dim1, dim2, dim3) outer bounding box for a pair-stacked core.

    The spreadsheet's column-label conventions differ between families
    (they follow each family's datasheet labeling, not a unified scheme):
      E / ETD / EFD: A = long axis, B = depth, C = height per half
                     → envelope = (A, B, 2*C)
      PQ:            A = long axis, B = height per half, C = depth
                     → envelope = (A, C, 2*B)
      RM:            J = across-flats (square footprint), B = height per half
                     → envelope = (J, J, 2*B)

    Returns None if required columns are missing.
    """
    col = {name: idx for idx, name in enumerate(headers)}
    try:
        if family in ("E", "ETD", "EFD"):
            A = row[col["A"]]
            B = row[col["B"]]
            C = row[col["C"]]
            return (A, B, 2 * C)
        if family == "PQ":
            A = row[col["A"]]
            B = row[col["B"]]
            C = row[col["C"]]
            return (A, C, 2 * B)
        if family == "RM":
            J = row[col["J"]]
            B = row[col["B"]]
            return (J, J, 2 * B)
    except (KeyError, TypeError):
        return None
    return None


def _load_envelopes() -> list[dict]:
    """Load all cores and their outer envelopes."""
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required. Install with: uv add openpyxl") from exc

    db_path = _find_db()
    wb = openpyxl.load_workbook(db_path, data_only=True, read_only=True)

    cores: list[dict] = []
    for family in wb.sheetnames:
        ws = wb[family]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = list(rows[0])
        for row in rows[1:]:
            if row[0] is None:
                continue
            env = _envelope_for_row(family, headers, row)
            if env is None:
                continue
            dims_sorted = tuple(sorted(env, reverse=True))
            cores.append({
                "name": str(row[0]),
                "family": family,
                "envelope_sorted_mm": dims_sorted,  # (largest, middle, smallest)
            })

    wb.close()
    return cores


def _fits(core_env: tuple[float, float, float], user_env: tuple[float, float, float]) -> bool:
    """Orientation-agnostic fit check — sort both, compare element-wise."""
    c = sorted(core_env, reverse=True)
    u = sorted(user_env, reverse=True)
    return all(c[i] <= u[i] for i in range(3))


class CoreEnvelopeTool(BaseTool):
    """Lookup or filter catalog cores by their outer bounding-box dimensions."""

    name = "core_envelope"
    description = (
        "Look up the outer bounding-box (physical envelope) of catalog cores (E, ETD, PQ, RM, EFD). "
        "Use this to check whether a core fits the user's dimension constraints or to filter "
        "candidates by envelope. Matching is orientation-agnostic — dimension labels "
        "(Height/Depth/Width) do not need to align; only the three numeric values do. "
        "Internally both sides are sorted descending and compared element-wise, so a "
        "'max height 15 mm' constraint matches any orientation where the core's tallest "
        "dimension is <= 15 mm. "
        "Modes: "
        "(a) pass `names` to get envelopes for specific cores; "
        "(b) pass `max_dims` (any 1-3 dims, in any order) to list all catalog cores that fit; "
        "(c) combine both to filter a named list by a constraint envelope."
    )
    parameters = {
        "type": "object",
        "properties": {
            "names": {
                "type": "array",
                "items": {"type": "string"},
                "description": "Optional list of catalog core names (e.g. ['PQ 20/20', 'ETD 34/17/11']).",
            },
            "max_dims": {
                "type": "array",
                "items": {"type": "number"},
                "description": (
                    "Optional 1-3 dimension constraints in mm, in any order. "
                    "If 1 value: treated as max-height only. "
                    "If 2 values: treated as footprint (length x width); height unconstrained. "
                    "If 3 values: full envelope. Orientation-agnostic."
                ),
            },
            "family": {
                "type": "string",
                "enum": ["E", "ETD", "PQ", "RM", "EFD"],
                "description": "Optional family filter.",
            },
        },
    }

    async def execute(self, project_dir: Path, **kwargs: Any) -> ToolResult:
        names = kwargs.get("names") or []
        max_dims = kwargs.get("max_dims") or []
        family_filter = kwargs.get("family")

        try:
            cores = _load_envelopes()
        except Exception as e:
            return ToolResult(tool_call_id="", content=f"Error loading DB: {e}", is_error=True)

        if family_filter:
            cores = [c for c in cores if c["family"] == family_filter]

        # Track which named cores were actually present in the DB so we can
        # surface "not in catalog" separately from "doesn't fit". Without this
        # the agent has been misreading absent-from-DB as no-fit.
        not_in_db: list[str] = []
        if names:
            def _key(s: str) -> str:
                return s.strip().upper().replace(" ", "")
            wanted_keys = {_key(n) for n in names}
            present_keys = {_key(c["name"]) for c in cores}
            not_in_db = sorted(_key(n) for n in names if _key(n) not in present_keys)
            cores = [c for c in cores if _key(c["name"]) in wanted_keys]

        out_of_envelope: list[dict] = []
        if max_dims:
            if len(max_dims) == 1:
                # Height-only: the largest of the three must be <= constraint
                constraint = (max_dims[0], 1e9, 1e9)
            elif len(max_dims) == 2:
                # Footprint only: two largest must fit; height unconstrained
                constraint = (max_dims[0], max_dims[1], 1e9)
            elif len(max_dims) == 3:
                constraint = tuple(max_dims)
            else:
                return ToolResult(
                    tool_call_id="",
                    content="Error: max_dims must have 1, 2, or 3 values.",
                    is_error=True,
                )
            fitting = [c for c in cores if _fits(c["envelope_sorted_mm"], constraint)]
            out_of_envelope = [c for c in cores if c not in fitting]
            cores = fitting

        if not cores and not not_in_db and not out_of_envelope:
            return ToolResult(
                tool_call_id="",
                content="No catalog cores match the given filters.",
            )

        lines = ["## Catalog core envelopes (outer bounding box, pair-stacked)"]
        lines.append("")
        lines.append("Dimensions shown sorted descending (largest → smallest). "
                     "Matching is orientation-agnostic.")
        lines.append("")

        if cores:
            lines.append("| Core | Family | Dim 1 (mm) | Dim 2 (mm) | Dim 3 (mm) |")
            lines.append("|---|---|---:|---:|---:|")
            for c in cores:
                d1, d2, d3 = c["envelope_sorted_mm"]
                lines.append(f"| {c['name']} | {c['family']} | {d1:.1f} | {d2:.1f} | {d3:.1f} |")
            lines.append("")

        if out_of_envelope:
            lines.append("### Cores in the catalog but exceeding the dimension constraint")
            lines.append("(these exist as standard parts, but their bounding box does NOT fit "
                         "the requested envelope in any orientation)")
            lines.append("")
            lines.append("| Core | Family | Dim 1 (mm) | Dim 2 (mm) | Dim 3 (mm) |")
            lines.append("|---|---|---:|---:|---:|")
            for c in out_of_envelope:
                d1, d2, d3 = c["envelope_sorted_mm"]
                lines.append(f"| {c['name']} | {c['family']} | {d1:.1f} | {d2:.1f} | {d3:.1f} |")
            lines.append("")

        if not_in_db:
            lines.append("### Cores NOT present in the envelope catalog")
            lines.append("(these names were requested but no entry exists in the geometry DB; "
                         "this is **not** a fit failure — dimensional fit is unknown for these)")
            lines.append("")
            for n in not_in_db:
                lines.append(f"- {n}")
            lines.append("")

        return ToolResult(tool_call_id="", content="\n".join(lines))
