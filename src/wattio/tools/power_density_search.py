"""Tool for searching the power density reference database."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from wattio.models import ToolResult
from wattio.tools.base import BaseTool

EXCEL_FILENAME = "POWER DENSITY.xlsx"

# Header names as they appear in the Excel file. Lookups are tolerant
# (whitespace + case insensitive) so harmless edits to the headers don't
# break the tool — but a missing required header raises an explicit error
# rather than silently picking the wrong column.
COL_SOURCE = "SOURCE"
COL_REFERENCE = "REFERENCE"
COL_MAG_TYPE = "MAGNETIC TYPE"
COL_TOPOLOGY = "TOPOLOGY"
COL_POWER = "POWER RATIO (kw) O INDUCTANCE"
COL_FREQUENCY = "SW FREQUENCY"
COL_RATIO = "RATIO"
COL_COOLING = "COOLING"
COL_CORE = "CORE"
COL_STACKS = "stacks"
COL_TECHNOLOGY = "TECHNOLOGY"
COL_VOL_EFF = "Volumen effective(l)"
COL_POWER_DENSITY = "DENSITY Ve (kw/l)"
COL_EFFICIENCY = "efficiency"
COL_TOTAL_LOSSES = "total magnetic losses"
COL_WINDING_LOSSES = "winding"
COL_CORE_LOSSES = "core"
COL_SIMULATOR = "simulator"

REQUIRED_COLS = (
    COL_SOURCE, COL_REFERENCE, COL_MAG_TYPE, COL_TOPOLOGY,
    COL_POWER, COL_FREQUENCY, COL_POWER_DENSITY,
)


def _normalize(s: object) -> str:
    # Whitespace-tolerant only — case is preserved because the spreadsheet
    # uses two columns that differ only by case ('CORE' = name, 'core' = losses).
    return str(s or "").strip()


def _build_header_index(headers: list) -> dict[str, int]:
    """Map normalized header → column index (first match wins)."""
    idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        key = _normalize(h)
        if key and key not in idx:
            idx[key] = i
    return idx


def _safe_float(value: object) -> float:
    """Coerce numeric cells; return 0.0 for blanks/strings/None."""
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _find_excel(project_dir: Path) -> Path:
    """Locate the power density Excel file."""
    for candidate in [
        project_dir / EXCEL_FILENAME,
        project_dir.parent / EXCEL_FILENAME,
        Path(__file__).parent.parent.parent.parent / EXCEL_FILENAME,
    ]:
        if candidate.is_file():
            return candidate
    raise FileNotFoundError(
        f"Cannot find '{EXCEL_FILENAME}'. Place it in the project directory."
    )


def _load_and_search(
    project_dir: Path,
    topology: str,
    magnetic_type: str,
) -> list[dict]:
    """Load Excel and find all matching designs."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required. Install with: uv add openpyxl")

    excel_path = _find_excel(project_dir)
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_idx = _build_header_index(headers)

    missing = [c for c in REQUIRED_COLS if _normalize(c) not in header_idx]
    if missing:
        wb.close()
        raise RuntimeError(
            f"Power-density spreadsheet is missing required column(s): "
            f"{missing}. Found headers: {[h for h in headers if h]}"
        )

    def cell(row: tuple, col_name: str, default: object = "") -> object:
        i = header_idx.get(_normalize(col_name))
        if i is None or i >= len(row):
            return default
        v = row[i]
        return default if v is None else v

    results: list[dict] = []
    topology_lower = topology.lower()
    mag_type_lower = magnetic_type.lower()

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_topology = str(cell(row, COL_TOPOLOGY)).lower()
        row_mag_type = str(cell(row, COL_MAG_TYPE)).lower()

        if topology_lower not in row_topology:
            continue
        if mag_type_lower not in row_mag_type:
            continue

        results.append({
            "source": cell(row, COL_SOURCE),
            "reference": cell(row, COL_REFERENCE),
            "topology": cell(row, COL_TOPOLOGY),
            "power_kw": round(_safe_float(cell(row, COL_POWER, 0)), 2),
            "frequency_khz": round(_safe_float(cell(row, COL_FREQUENCY, 0)), 1),
            "turns_ratio": cell(row, COL_RATIO),
            "cooling": cell(row, COL_COOLING),
            "core": cell(row, COL_CORE),
            "stacks": cell(row, COL_STACKS, 1),
            "technology": str(cell(row, COL_TECHNOLOGY)).strip(),
            "vol_effective_l": round(_safe_float(cell(row, COL_VOL_EFF, 0)), 6),
            "power_density_kw_l": round(_safe_float(cell(row, COL_POWER_DENSITY, 0)), 1),
            "efficiency_pct": cell(row, COL_EFFICIENCY),
            "total_losses_w": cell(row, COL_TOTAL_LOSSES),
            "winding_losses_w": cell(row, COL_WINDING_LOSSES),
            "core_losses_w": cell(row, COL_CORE_LOSSES),
            "simulator_link": cell(row, COL_SIMULATOR),
        })

    wb.close()

    results.sort(key=lambda r: r["power_density_kw_l"], reverse=True)
    return results


class PowerDensitySearchTool(BaseTool):
    """Search the power density reference database for similar designs."""

    name = "power_density_search"
    description = (
        "Search the power density reference database (Excel) to find real-world designs "
        "with similar topology and magnetic type. Returns power density benchmarks from "
        "academic papers and application notes, including core used, frequency, cooling, "
        "losses breakdown, and efficiency. Use this in Step 4 to give the engineer a "
        "reference point for achievable power density before starting viability screening."
    )
    parameters = {
        "type": "object",
        "properties": {
            "topology": {
                "type": "string",
                "description": "Converter topology to search for (e.g., 'PSFB', 'LLC', 'DAB').",
            },
            "magnetic_type": {
                "type": "string",
                "enum": ["transformer", "inductor"],
                "description": "Type of magnetic component.",
            },
        },
        "required": ["topology", "magnetic_type"],
    }

    async def execute(self, project_dir: Path, **kwargs: Any) -> ToolResult:
        topology = kwargs.get("topology", "")
        magnetic_type = kwargs.get("magnetic_type", "transformer")

        if not topology:
            return ToolResult(
                tool_call_id="",
                content="Error: topology is required.",
                is_error=True,
            )

        try:
            results = _load_and_search(
                project_dir=project_dir,
                topology=topology,
                magnetic_type=magnetic_type,
            )
        except FileNotFoundError as e:
            return ToolResult(tool_call_id="", content=f"Error: {e}", is_error=True)
        except Exception as e:
            return ToolResult(
                tool_call_id="", content=f"Error searching database: {e}", is_error=True
            )

        if not results:
            return ToolResult(
                tool_call_id="",
                content=(
                    f"No reference designs found for topology '{topology}' "
                    f"with magnetic type '{magnetic_type}'. "
                    f"The database may not have entries for this combination yet."
                ),
            )

        # Format results
        lines = [
            f"## Power Density Reference Designs",
            f"**Search:** {magnetic_type} designs for {topology} topology",
            f"**Found:** {len(results)} reference design(s)",
            "",
        ]

        for i, r in enumerate(results, 1):
            power_w = r["power_kw"] * 1000
            lines.append(
                f"### {i}. {r['source']}: {r['topology']} {power_w:.0f} W @ "
                f"{r['frequency_khz']:.0f} kHz"
            )
            lines.append("")
            lines.append(f"| Parameter | Value |")
            lines.append(f"|-----------|-------|")
            lines.append(f"| Source | {r['source']} |")
            lines.append(f"| Reference | {r['reference']} |")
            lines.append(f"| Power | {power_w:.0f} W ({r['power_kw']} kW) |")
            lines.append(f"| Frequency | {r['frequency_khz']:.0f} kHz |")
            lines.append(f"| Turns ratio | {r['turns_ratio']} |")
            lines.append(f"| Cooling | {r['cooling']} |")
            lines.append(f"| Core | {r['core']} (×{r['stacks']}) |")
            lines.append(f"| Technology | {r['technology']} |")
            lines.append(f"| Effective volume | {r['vol_effective_l']} L |")
            lines.append(f"| **Power density** | **{r['power_density_kw_l']:.1f} kW/L** |")
            lines.append(f"| Efficiency | {r['efficiency_pct']}% |")
            lines.append(f"| Total magnetic losses | {r['total_losses_w']} W |")
            lines.append(f"| Winding losses | {r['winding_losses_w']} W |")
            lines.append(f"| Core losses | {r['core_losses_w']} W |")
            if r["simulator_link"]:
                lines.append(f"| Frenetic simulation | {r['simulator_link']} |")
            lines.append("")

        lines.append(
            "**How to use:** Compare your design's target power and frequency with these references. "
            "The power density (kW/L) gives you a benchmark for what's achievable with similar "
            "topologies. If a reference design uses a similar power level and frequency, its core "
            "choice and loss distribution can inform your own design targets."
        )

        return ToolResult(
            tool_call_id="",
            content="\n".join(lines),
        )
